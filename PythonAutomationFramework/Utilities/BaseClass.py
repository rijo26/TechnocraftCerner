import os
import random
import string
import time

import allure
import pytest
import inspect
import logging
import openpyxl
from allure_commons.types import AttachmentType
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait


@pytest.mark.usefixtures('setup')
class BaseClass:

    def enterDetailsOnLoginScreen(self):
        testCaseName = inspect.stack()[2][3]
        wb = openpyxl.load_workbook('..\\TestDataSetUp\\Excels\\TestData.xlsx')
        sheet = wb['Sheet1']
        max_row1 = sheet.max_row
        userName = None
        password = None
        organization = None
        for a in range(1, max_row1 + 1):
            if sheet.cell(row=a, column=1).value == testCaseName:
                userName = sheet.cell(row=a, column=2).value
                password = sheet.cell(row=a, column=3).value
                organization = sheet.cell(row=a, column=4).value
                break

        self.driver.find_element_by_id('username').send_keys(userName)
        self.driver.find_element_by_id('password').send_keys(password)
        time.sleep(2)
        self.selectOptionByText(self.driver.find_element_by_id('post_o_org_id'), organization)
        return

    def verifyWaitUntilForLinkText(self, element):
        ele = WebDriverWait(self.driver, 320).until(
            EC.presence_of_element_located((By.LINK_TEXT, element))
        )

    def verifyWaitUntilForTagName(self, element):
        ele = WebDriverWait(self.driver, 320).until(
            EC.presence_of_element_located((By.TAG_NAME, element))
        )

    def selectOptionByText(self, locator, text):
        sel = Select(locator)
        sel.select_by_visible_text(text)

    def getLogger(self, method, function):
        loggerName1 = method
        loggerName2 = function
        loggerName = loggerName1 + " -> " + loggerName2
        logger = logging.getLogger(
            loggerName)  # __name__ returns the testcase name. if not given this will return the name root

        file = logging.FileHandler('..\\LogFolder\\logfile.log')  # This will tell where to write the logs
        # Format as to how to print, what info to be printed in the logger file
        formatter = logging.Formatter("%(asctime)s : %(levelname)s : %(name)s : %(message)s")
        file.setFormatter(formatter)  # This is to send the format along with the file handler
        if logger.hasHandlers():
            logger.handlers.clear()
        logger.addHandler(file)  # pass file handler object
        logger.setLevel(logging.INFO)
        return logger

    def grabScreenShot(self, name):
        my_path = os.path.abspath(os.path.dirname(__file__))
        loggerName1 = inspect.stack()[2][3]
        loggerName2 = inspect.stack()[1][3]
        names = name + "_" + loggerName2
        timestr = time.strftime("%d-%b-%Y")
        screenShotPath = os.path.join(my_path, "..//Screenshots//" + loggerName1 + "//")
        isExist = os.path.exists(screenShotPath)
        if isExist:
            self.driver.save_screenshot(
                screenShotPath + loggerName1 + "_" + names + "_" + timestr + "_" + UUID() + ".png")
        else:
            createFolder(screenShotPath)
            self.driver.save_screenshot(
                screenShotPath + loggerName1 + "_" + names + "_" + timestr + "_" + UUID() + ".png")
        return

    def reportLogs(self, method, loggerINFO, step, attachmentName):
        #   This will return the method from which this statement is called
        log = self.getLogger(method, inspect.stack()[1][3])
        log.info(loggerINFO)
        self.allureDetails(step, attachmentName)
        return

    def allureDetails(self, step, attachmentName):
        allure.step(step)
        allure.attach(self.driver.get_screenshot_as_png(), name=attachmentName, attachment_type=AttachmentType.PNG)
        return


def UUID(size=8, chars=string.ascii_uppercase + string.digits):
    return ''.join(random.choice(chars) for _ in range(size))


def createFolder(path):
    try:
        os.mkdir(path)
    except OSError:
        print("Creation of directory %s failed" % path)
    else:
        print("Creation of directory successful")
